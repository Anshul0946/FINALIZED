"""
excel_handler.py
Handles all Excel workbook operations: image extraction, template scanning, value mapping.
"""

import os
import io
import re
import json
import openpyxl
from PIL import Image
from pathlib import Path
from typing import List


def get_sector_from_col(col_index: int) -> str:
    """Determine sector based on column index"""
    if 0 <= col_index < 4:
        return "alpha"
    if 4 <= col_index < 8:
        return "beta"
    if 8 <= col_index < 12:
        return "gamma"
    if 12 <= col_index < 18:
        return "voicetest"
    return "unknown"


def extract_images_from_excel(xlsx_path: str, output_folder: str, 
                              log_callback) -> List[str]:
    """Extract all images from Excel workbook and save to disk"""
    log_callback(f"[EXCEL] Analyzing template file: {xlsx_path}")
    
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        sheet = wb.active
    except Exception as e:
        log_callback(f"[EXCEL ERROR] Could not open/read Excel file: {e}")
        return []
    
    images = getattr(sheet, "_images", [])
    if not images:
        log_callback("[EXCEL] No images found in workbook.")
        return []
    
    os.makedirs(output_folder, exist_ok=True)
    images_with_locations = []
    
    for image in images:
        try:
            row = image.anchor._from.row + 1
            col = image.anchor._from.col
        except Exception:
            row, col = 0, 0
        images_with_locations.append({"image": image, "row": row, "col": col})
    
    images_sorted = sorted(images_with_locations, key=lambda i: (i["row"], i["col"]))
    saved_paths = []
    counters = {"alpha": 0, "beta": 0, "gamma": 0, "voicetest": 0, "unknown": 0}
    
    log_callback(f"[EXCEL] Found {len(images_sorted)} images. Extracting...")
    
    for itm in images_sorted:
        sector = get_sector_from_col(itm["col"])
        counters[sector] += 1
        filename = f"{sector}_image_{counters[sector]}.png"
        out_path = os.path.join(output_folder, filename)
        
        try:
            img_data = itm["image"]._data()
            pil = Image.open(io.BytesIO(img_data))
            pil.save(out_path, "PNG")
            saved_paths.append(out_path)
            
            try:
                loc = f"{openpyxl.utils.get_column_letter(itm['col']+1)}{itm['row']}"
            except Exception:
                loc = ""
            log_callback(f"[EXCEL] Saved {filename} at {loc}")
        except Exception as e:
            log_callback(f"[EXCEL ERROR] Failed to save {filename}: {e}")
    
    return saved_paths


def scan_bold_red_expressions(xlsx_path: str, log_callback) -> List[tuple]:
    """Scan workbook for bold+red cells and extract expressions"""
    log_callback("[EXCEL] Scanning for bold+red expressions...")
    
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        sheet = wb.active
    except Exception as e:
        log_callback(f"[EXCEL ERROR] Could not open workbook: {e}")
        return []
    
    cells_to_process = []
    
    def _font_is_strict_red(font):
        if not font:
            return False
        if not getattr(font, "bold", False):
            return False
        col = getattr(font, "color", None)
        if col is None:
            return False
        rgb = getattr(col, "rgb", None)
        if not rgb:
            return False
        up = str(rgb).upper()
        return up[-6:] == "FF0000"
    
    def _normalize_expr(raw: str) -> str:
        s = raw.strip()
        if (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'")):
            s = s[1:-1].strip()
        return s
    
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=16):
        for cell in row:
            val = cell.value
            if not val or not isinstance(val, str):
                continue
            font = cell.font
            if not font:
                continue
            
            if _font_is_strict_red(font):
                expr = _normalize_expr(val)
                if expr:
                    cells_to_process.append((cell, expr))
    
    log_callback(f"[EXCEL] Found {len(cells_to_process)} bold+red expressions")
    return cells_to_process


def map_values_to_template(xlsx_path: str, cells_to_process: List[tuple], 
                          allowed_vars: dict, resolve_func, log_callback) -> str:
    """Map resolved values back to Excel template"""
    log_callback("[EXCEL] Mapping values to template...")
    
    try:
        wb_edit = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        log_callback(f"[EXCEL ERROR] Could not open workbook for mapping: {e}")
        return xlsx_path
    
    def _to_number_convert(v):
        try:
            if v is None:
                return None
            if isinstance(v, (int, float)):
                return v
            if isinstance(v, bool):
                return None
            s = str(v).strip()
            s_clean = s.replace(",", "")
            if s_clean == "":
                return None
            if re.fullmatch(r"[-+]?\d+", s_clean):
                return int(s_clean)
            if re.fullmatch(r"[-+]?\d*\.\d+", s_clean):
                return float(s_clean)
            return None
        except Exception:
            return None
    
    for cell_obj, expr in cells_to_process:
        resolved = resolve_func(expr, allowed_vars)
        
        if resolved is None:
            cell_obj.value = "NULL"
        else:
            if isinstance(resolved, str):
                conv = _to_number_convert(resolved)
                if conv is not None:
                    cell_obj.value = conv
                else:
                    cell_obj.value = resolved
            elif isinstance(resolved, (int, float)):
                cell_obj.value = resolved
            elif isinstance(resolved, (dict, list)):
                try:
                    cell_obj.value = json.dumps(resolved)
                except Exception:
                    cell_obj.value = str(resolved)
            else:
                cell_obj.value = str(resolved)
    
    try:
        wb_edit.save(xlsx_path)
        log_callback(f"[EXCEL] ✓ Workbook saved: {xlsx_path}")
    except Exception as e:
        log_callback(f"[EXCEL ERROR] Failed to save workbook: {e}")
    
    return xlsx_path
